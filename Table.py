import streamlit as st
import pandas as pd
import os
import io

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ----------------------------------------------------------------
# 1) Helper function: add_entity_info
# ----------------------------------------------------------------
def add_entity_info(ws, entity_info, start_row):
    """
    Example function to write multi-line 'entity_info' to the worksheet at 'start_row'.
    Each line can be styled or merged as needed.
    """
    for i, line in enumerate(entity_info.split('\n'), start=1):
        cell = ws.cell(row=start_row + i - 1, column=1)
        cell.value = line
        cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000")
        )
        # Example: special formatting if the line starts with "Source:", "Entity:", etc.
        if line.startswith('Entity:'):
            cell.font = Font(bold=True, color="000000", name="Gill Sans")
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        if line.startswith('Source:'):
            cell.font = Font(color="000000", name="Gill Sans")
        if line.startswith('Time Period of analysis:'):
            cell.font = Font(color="000000", name="Gill Sans")

# ----------------------------------------------------------------
# 2) Helper function: add_styling_to_worksheet
# ----------------------------------------------------------------
def add_styling_to_worksheet(ws, df, start_row, comment):
    """
    Example function to write a comment/title row for a DataFrame
    and then write the DataFrame below it with basic styling.
    """
    # 2a) Write the comment/title in a merged cell
    cell = ws.cell(row=start_row, column=1)
    cell.value = comment
    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    cell.font = Font(color="000000", bold=True, name="Gill Sans")
    cell.alignment = Alignment(horizontal='center')
    ws.merge_cells(
        start_row=start_row, 
        start_column=1, 
        end_row=start_row, 
        end_column=len(df.columns)  # merge across the number of columns in df
    )

    # 2b) Write DF column headers in the next row
    header_row = start_row + 1
    for col_idx, col_name in enumerate(df.columns, start=1):
        hdr_cell = ws.cell(row=header_row, column=col_idx)
        hdr_cell.value = col_name
        hdr_cell.font = Font(bold=True, name="Gill Sans")
        hdr_cell.alignment = Alignment(horizontal='center')
        hdr_cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000")
        )

    # 2c) Write DF rows
    data_start_row = header_row + 1
    for row_idx, row_data in enumerate(df.values, start=data_start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell_ = ws.cell(row=row_idx, column=col_idx)
            cell_.value = value
            cell_.font = Font(name="Gill Sans")
            cell_.alignment = Alignment(horizontal='center')
            cell_.border = Border(
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000")
            )

# ----------------------------------------------------------------
# 3) Main function: multiple_dfs (returns bytes in memory)
# ----------------------------------------------------------------
def multiple_dfs(df_list, sheet_name, comments, entity_info):
    """
    Build an Excel workbook with the given list of DataFrames + comments,
    using openpyxl for formatting, and return the resulting bytes (no disk write).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Start row for the first block
    current_row = 1

    # (Optional) Add entity info in the first few rows
    add_entity_info(ws, entity_info, current_row)
    # Move down a bit after writing entity info
    current_row += 6

    # Loop over each (DataFrame, comment) pair
    for each_df, comment in zip(df_list, comments):
        # Add styling / title row
        add_styling_to_worksheet(ws, each_df, current_row, comment)
        # Increase current_row by the number of DF rows + 2 (header + comment row, etc.)
        # We'll approximate how many lines we used in add_styling_to_worksheet
        current_row += (len(each_df) + 2)  # or add more if needed

        # Optionally add some spacing before the next table
        current_row += 2

    # Save to an in-memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()  # Return the raw Excel file bytes

# ----------------------------------------------------------------
# Streamlit App
# ----------------------------------------------------------------
st.title("Online News Data Extraction")

# 4) File uploader for initial multiple Excel files
uploaded_files = st.file_uploader(
    "Please Upload your file (in .xlsx format)",
    type=["xlsx"],
    accept_multiple_files=True
)

if uploaded_files:
    st.success("Files uploaded successfully!")

    # 5) Read and combine the uploads
    dfs = []
    for uploaded_file in uploaded_files:
        filename_no_ext = os.path.splitext(uploaded_file.name)[0]
        entity_name = filename_no_ext.split(" - ")[0]

        df_ = pd.read_excel(uploaded_file)
        df_["Entity"] = entity_name
        dfs.append(df_)

    combined_df = pd.concat(dfs, axis=0, ignore_index=True)

    # Just an example utility to quickly download combined_df
    def to_excel_bytes(df, sheet_name):
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        out.seek(0)
        return out.getvalue()

    # Download button for combined_df
    excel_data = to_excel_bytes(combined_df, "Combined Data")
    st.download_button(
        label="Download Combined Data as Excel",
        data=excel_data,
        file_name="combined_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # 6) Optional: second upload for an edited file
    edited_file = st.file_uploader("Upload an edited Excel file (.xlsx)", type="xlsx")
    if edited_file is not None:
        df = pd.read_excel(edited_file)
        st.header("Extract Dataframes")
        Entity_SOV = pd.crosstab(df['Entity'], columns='News Count', values=df['Entity'], aggfunc='count', margins=True, margins_name='Total').round(0)
        Entity_SOV.drop(columns='Total', inplace=True, errors='ignore')
        df_no_total = Entity_SOV.loc[Entity_SOV.index != 'Total']
        sum_no_total = df_no_total['News Count'].sum()
        Entity_SOV.loc[df_no_total.index, '%'] = (df_no_total['News Count'] / sum_no_total * 100).round(2)
        Entity_SOV.loc['Total', '%'] = 100
        Entity_SOV = pd.DataFrame(Entity_SOV.to_records())
        df['Date'] = pd.to_datetime(df['Date']).dt.normalize()
        sov_dt = pd.crosstab((df['Date'].dt.to_period('M')), df['Entity'], margins=True, margins_name='Total')
        Publication = pd.crosstab(df['Publication Name'], df['Entity'])
        Publication['Total'] = Publication.sum(axis=1)
        Publication = Publication.sort_values(by='Total', ascending=False).round()
        Publication.loc['GrandTotal'] = Publication.sum(numeric_only=True, axis=0)
        Publication = pd.DataFrame(Publication.to_records())
        df1= df.copy()
        df1['Journalist'] = df1['Journalist'].str.split(',')
        df1 = df1.explode('Journalist')
        jr_tab = pd.crosstab(df1['Journalist'], df1['Entity'])
        jr_tab = jr_tab.reset_index(level=0)
        newdata = df1[['Journalist', 'Publication Name']]
        Journalist_Table = pd.merge(jr_tab, newdata, how='inner', left_on=['Journalist'], right_on=['Journalist'])
        Journalist_Table.drop_duplicates(subset=['Journalist'], keep='first', inplace=True)
        valid_columns = Journalist_Table.select_dtypes(include='number').columns
        Journalist_Table['Total'] = Journalist_Table[valid_columns].sum(axis=1)
        Jour_table = Journalist_Table.sort_values('Total', ascending=False).round()
        bn_row = Jour_table.loc[Jour_table['Journalist'] == 'Bureau News']
        Jour_table = Jour_table[Jour_table['Journalist'] != 'Bureau News']
        Jour_table = pd.concat([Jour_table, bn_row], ignore_index=True)
#         Jour_table = Journalist_Table.reset_index(drop=True)
        Jour_table.loc['GrandTotal'] = Jour_table.sum(numeric_only=True, axis=0)
        columns_to_convert = Jour_table.columns.difference(['Journalist', 'Publication Name'])
        Jour_table[columns_to_convert] = Jour_table[columns_to_convert].astype(int)
        Jour_table.insert(1, 'Publication Name', Jour_table.pop('Publication Name'))
        Jour_table.loc['Total'] = Jour_table.sum(numeric_only=True, axis=0)
        # pubs_table1['% '] = pubs_table1['% '].astype(int)
        Jour_table = pd.DataFrame(Jour_table.to_records())

        PT_Entity = pd.crosstab(df['Publication Type'], df['Entity'])
        PT_Entity['Total'] = PT_Entity.sum(axis=1)
        PType_Entity = PT_Entity.sort_values('Total', ascending=False).round()
        PType_Entity.loc['GrandTotal'] = PType_Entity.sum(numeric_only=True, axis=0)
        PType_Entity = pd.DataFrame(PType_Entity.to_records())
        crosstab = pd.crosstab(df['Journalist'], df['Entity'], margins=True, margins_name='Total')
        crosstab = pd.DataFrame(crosstab.to_records())
        client_cols = [col for col in crosstab.columns if col.startswith('Client-')]
        competitor_cols = [col for col in crosstab.columns if col not in ['Journalist', 'Total'] and not col.startswith('Client-')]
        mask_client_positive = (crosstab[client_cols].sum(axis=1) != 0) if client_cols else False
        mask_competitors_zero = (crosstab[competitor_cols].sum(axis=1) == 0) if competitor_cols else True
        j_abt_client = crosstab[mask_client_positive & mask_competitors_zero]
        j_abt_client = j_abt_client.sort_values(by='Total', ascending=False)
        crosstab1 = pd.crosstab(df['Journalist'], df['Entity'], margins=True, margins_name='Total')
        crosstab1 = pd.DataFrame(crosstab1.to_records())
        mask_client_zero = (crosstab1[client_cols].sum(axis=1) == 0) if client_cols else True
        mask_competitor_positive = (crosstab1[competitor_cols].sum(axis=1) != 0) if competitor_cols else False
        j_abt_comp = crosstab1[mask_client_zero & mask_competitor_positive]
        j_abt_comp = j_abt_comp.sort_values(by='Total', ascending=False)
        df_dict = {
        "Entity_SOV": Entity_SOV,
        "M-O-M SOV Table": sov_dt,
        "Publication Table": Publication,
        "Journalist Table":  Jour_table,
        'Pub Type and Entity Table': PType_Entity,
        'Journalist writing on Client and not on Comp': j_abt_client,
        'Journalist writing on Comp and not on Client': j_abt_comp}
        selected_option = st.selectbox("Pick a DataFrame to view", list(df_dict.keys()))
        st.dataframe(df_dict[selected_option])
        dfs_for_export = list(df_dict.values())
        comments = [
        'SOV Table', 
        'Month-on-Month Table', 
        'Publication Table', 
        'Journalist Table All',
        'Pub Type and Entity Table',
        'Journalist writing on Client and not on Comp', 
        'Journalist writing on Comp and not on Client']
        entity_info = """Entity:
Time Period of analysis: 19th April 2023 to 18th April 2024
Source: (Online) Meltwater, Select 100 online publications...
News search: All Articles: entity mentioned at least once
"""
        excel_bytes = multiple_dfs(
        df_list=dfs_for_export,
        sheet_name='Results',
        comments=comments,
        entity_info=entity_info)
        st.download_button(
        label="Download All Dataframes (XLSX)",
        data=excel_bytes,
        file_name="All_Dataframes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" )



    